from django.shortcuts import render, redirect, get_object_or_404
from datetime import date, timedelta
from .models import Sala, Osoba, Sprzet, Dozymetr
from .forms import SalaForm, OsobaForm, SprzetForm, DozymetrForm
from django.db.models import Prefetch, Count, Q
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.http import HttpResponse
from django.views import View
import openpyxl
from io import BytesIO
from django.http import HttpResponseBadRequest
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
from django.utils.datastructures import MultiValueDictKeyError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import json


def startowa_strona(request):
    context = {
        'message': "Co chcesz wybrać Drogi Użytkowniku?",
        'selected': None,
    }

    if request.method == 'POST':
        context['selected'] = request.POST.get('wybor', 'Nic nie wybrano')

    sala_counts = (
        Sala.objects
        .annotate(prom_count=Count('sprzet', filter=Q(sprzet__promieniujacy=True)))
        .order_by('-prom_count')
    )

    context['chart_labels'] = json.dumps([s.nr_sali or 'Brak' for s in sala_counts])
    context['chart_data'] = json.dumps([s.prom_count for s in sala_counts])

    return render(request, 'start.html', context)

# tworzenie sali
def sala_list(request):
    sale_all = Sala.objects.all()
    paginator = Paginator(sale_all, 4)  

    page_number = request.GET.get('page')
    sale = paginator.get_page(page_number)

    sprzety = Sprzet.objects.all()
    dozymetry = Dozymetr.objects.select_related('sala')

    sprzety_by_sala = {}
    dozymetry_by_sala = {}

    for sprzet in sprzety:
        sprzety_by_sala.setdefault(sprzet.sala, []).append(sprzet)

    for dozymetr in dozymetry:
        dozymetry_by_sala.setdefault(dozymetr.sala, []).append(dozymetr)

    return render(request, 'sala_list.html', {
        'sale': sale,
        'sprzety_by_sala': sprzety_by_sala,
        'dozymetry_by_sala': dozymetry_by_sala,
    })

def sala_create(request):
    if request.method == 'POST':
        form = SalaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('sala_list')
    else:
        form = SalaForm()
    return render(request, 'sala_form.html', {'form': form})

# tworzenie osoby
def osoba_list(request):
    osoby_list = Osoba.objects.all()
    paginator = Paginator(osoby_list, 4)
    page = request.GET.get('page')
    try:
        osoby = paginator.get_page(page)
    except PageNotAnInteger:
        osoby = paginator.get_page(1)
    except EmptyPage:
        osoby = paginator.get_page(paginator.num_pages)

    dozymetry = Dozymetr.objects.select_related('osoba')
    dozymetry_by_osoba = {}

    for d in dozymetry:
        osoba = d.osoba
        if osoba not in dozymetry_by_osoba:
            dozymetry_by_osoba[osoba] = []

        if d.data_ostatniej_kontroli:
            next_calibration = d.data_ostatniej_kontroli + timedelta(days=90)
            days_left = (next_calibration - date.today()).days
            needs_alert = days_left < 14
        else:
            next_calibration = None
            days_left = None
            needs_alert = True

        dozymetry_by_osoba[osoba].append({
            'id': d.IDdozymetru,
            'status': d.status,
            'status_display': d.get_status_display(),
            'data_kontroli': d.data_ostatniej_kontroli,
            'days_left': days_left,
            'needs_alert': needs_alert,
        })

    return render(request, 'osoba_list.html', {
        'osoby': osoby,
        'dozymetry_by_osoba': dozymetry_by_osoba,
    })

def osoba_create(request):
    if request.method == 'POST':
        form = OsobaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('osoba_list')
    else:
        form = OsobaForm()
    return render(request, 'osoba_form.html', {'form': form})

# tworzenie sprzętu
def sprzet_list(request):
    sprzety_list = Sprzet.objects.select_related('sala').all()

    sala = request.GET.get('sala')
    promieniujacy = request.GET.get('promieniujacy')

    if sala:
        sprzety_list = sprzety_list.filter(sala__nr_sali__icontains=sala)

    if promieniujacy == 'true':
        sprzety_list = sprzety_list.filter(promieniujacy=True)
    elif promieniujacy == 'false':
        sprzety_list = sprzety_list.filter(promieniujacy=False)

    paginator = Paginator(sprzety_list, 10)
    page = request.GET.get('page')

    try:
        sprzety = paginator.page(page)
    except PageNotAnInteger:
        sprzety = paginator.page(1)
    except EmptyPage:
        sprzety = paginator.page(paginator.num_pages)

    return render(request, 'sprzet_list.html', {
        'sprzety': sprzety,
        'sale': Sala.objects.all(),  
        'request': request 
    })

def sprzet_create(request):
    if request.method == 'POST':
        form = SprzetForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('sprzet_list')
    else:
        form = SprzetForm()
    return render(request, 'sprzet_form.html', {'form': form})

# tworzenie dozymetrów
def dozymetr_list(request):
    dozymetry_qs = Dozymetr.objects.select_related('sala', 'osoba').all()

    paginator = Paginator(dozymetry_qs, 4)
    page_number = request.GET.get('page')

    try:
        dozymetry = paginator.page(page_number)
    except PageNotAnInteger:
        dozymetry = paginator.page(1)
    except EmptyPage:
        dozymetry = paginator.page(paginator.num_pages)

    for d in dozymetry:
        if d.data_ostatniej_kontroli:
            next_date = d.data_ostatniej_kontroli + timedelta(days=90)
            days_left = (next_date - date.today()).days
            needs_alert = days_left < 14
        else:
            days_left = None
            needs_alert = True
        d.days_left = days_left
        d.needs_alert = needs_alert

    return render(request, 'dozymetr_list.html', {
        'dozymetry': dozymetry
    })

def dozymetr_create(request):
    if request.method == 'POST':
        form = DozymetrForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dozymetr_list')
    else:
        form = DozymetrForm()
    return render(request, 'dozymetr_form.html', {'form': form})

def home_view(request):
    return render(request, 'home.html')

# usuwanie
@require_POST
def sala_delete(request, pk):
    sala = get_object_or_404(Sala, pk=pk)
    try:
        sala.delete()
        messages.success(request, 'Sala została pomyślnie usunięta.')
    except Exception as e:
        messages.error(request, f'Wystąpił błąd podczas usuwania sali: {e}')
    return redirect('sala_list')

@require_POST
def osoba_delete(request, pk):
    osoba = get_object_or_404(Osoba, pk=pk)
    try:
        osoba.delete()
        messages.success(request, 'Osoba została pomyślnie usunięta.')
    except Exception as e:
        messages.error(request, f'Wystąpił błąd podczas usuwania osoby: {e}')
    return redirect('osoba_list')

@require_POST
def sprzet_delete(request, pk):
    sprzet = get_object_or_404(Sprzet, pk=pk)
    try:
        sprzet.delete()
        messages.success(request, 'Sprzęt został pomyślnie usunięty.')
    except Exception as e:
        messages.error(request, f'Wystąpił błąd podczas usuwania sprzętu: {e}')
    return redirect('sprzet_list')

@require_POST
def dozymetr_delete(request, pk):
    dozymetr = get_object_or_404(Dozymetr, pk=pk)
    try:
        dozymetr.delete()
        messages.success(request, 'Dozymetr został pomyślnie usunięty.')
    except Exception as e:
        messages.error(request, f'Wystąpił błąd podczas usuwania dozymetru: {e}')
    return redirect('dozymetr_list')

class ExportXLSXView(View):
    models = {
        'sale': {
            'model': Sala,
            'fields': ['nr_sali', 'oddzial', 'sprzety', 'dozymetry'],  
            'filename': 'lista_sal.xlsx'
        },
        'osoby': {
            'model': Osoba,
            'fields': ['imie', 'nazwisko', 'dozymetry'],
            'filename': 'lista_osob.xlsx'
        },
        'sprzety': {
            'model': Sprzet,
            'fields': ['nazwa','producent_model','typ_sprzetu','data_ostatniego_przegladu','promieniujacy','status',
            'numer_seryjny','rok_zakupu','numer_kontaktowy_serwis','numer_ref','sala',],
            'filename': 'lista_sprzetu.xlsx'
        },
        'dozymetry': {
            'model': Dozymetr,
            'fields': ['IDdozymetru', 'typ', 'status', 'data_ostatniej_kontroli', 'sala', 'osoba'],
            'filename': 'lista_dozymetrow.xlsx'
        }
    }

    def get(self, request, model_name):
        if model_name not in self.models:
            return HttpResponse("Nieprawidłowy model", status=400)

        config = self.models[model_name]
        queryset = config['model'].objects.all()
        
        return self.generate_xlsx(queryset, config['fields'], config['filename'])

    def generate_xlsx(self, queryset, fields, filename):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Dane"

        for col_num, field in enumerate(fields, 1):
            worksheet.cell(row=1, column=col_num, value=field)

        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(fields, 1):
                value = getattr(obj, field)
                if hasattr(value, '__dict__'):
                    value = str(value)
                worksheet.cell(row=row_num, column=col_num, value=value)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class ImportXLSXView(View):

    required = {
        'sale':   ['nr_sali', 'oddzial'],
        'osoby':  ['imie', 'nazwisko'],
        'sprzety': ['model', 'producent', 'typ_sprzetu',
                    'data_ostatniego_przegladu', 'promieniujacy',
                    'status', 'numer_seryjny', 'rok_zakupu',
                    'numer_kontaktowy_serwis', 'sala'],
        'dozymetry': ['IDdozymetru','data_ostatniej_kontroli','status','typ','sala','osoba'],
    }

    def post(self, request, model_name):
        if model_name not in self.required:
            return HttpResponseBadRequest("Nieobsługiwany model do importu.")

        f = request.FILES.get('file')
        if not f:
            return HttpResponseBadRequest("Nie przesłano pliku.")

        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            req = self.required[model_name]
            missing_cols = set(req) - set(headers)
            if missing_cols:
                return HttpResponseBadRequest(
                    f"Brak kolumn: {', '.join(missing_cols)}"
                )

            created = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                if model_name == 'sale':
                    Sala.objects.get_or_create(
                        nr_sali=str(data['nr_sali']).strip(),
                        oddzial=str(data['oddzial']).strip()
                    )
                elif model_name == 'osoby':
                    Osoba.objects.get_or_create(
                        imie=str(data['imie']).strip(),
                        nazwisko=str(data['nazwisko']).strip()
                    )
                elif model_name == 'sprzety':
                    sala = Sala.objects.get(nr_sali=str(data['sala']).strip())
                    Sprzet.objects.create(
                        model=data['model'],
                        producent=data['producent'],
                        typ_sprzetu=data['typ_sprzetu'],
                        data_ostatniego_przegladu=data.get('data_ostatniego_przegladu'),
                        promieniujacy=bool(data['promieniujacy']),
                        status=data['status'],
                        numer_seryjny=str(data['numer_seryjny']),
                        rok_zakupu=str(data['rok_zakupu']),
                        numer_kontaktowy_serwis=data.get('numer_kontaktowy_serwis'),
                        sala=sala
                    )
                elif model_name == 'dozymetry':
                    sala = Sala.objects.get(nr_sali=str(data['sala']).strip())
                    
                    osoba = None
                    osoba_raw = str(data['osoba']).strip()
                    if osoba_raw and osoba_raw.lower() != 'none':
                        parts = osoba_raw.split()
                        if len(parts) >= 2:
                            imie = parts[0].strip()
                            nazwisko = ' '.join(parts[1:]).strip()
                            osoba = Osoba.objects.filter(imie__iexact=imie, nazwisko__iexact=nazwisko).first()
                        else:
                            raise ValueError(f"Nieprawidłowy format osoby: '{osoba_raw}' (oczekiwano 'Imie Nazwisko')")
                        
                    Dozymetr.objects.create(
                    IDdozymetru=str(data['IDdozymetru']).strip(),
                    typ=data['typ'].strip(),
                    status=data['status'].strip(),
                    data_ostatniej_kontroli=data['data_ostatniej_kontroli'],
                    sala=sala,
                    osoba=osoba    
                    )
                created += 1

            messages.success(request, f"Pomyślnie zaimportowano {created} rekordów.")

            redirect_map = {
                'sale': 'sala_list',
                'osoby': 'osoba_list',
                'sprzety': 'sprzet_list',
                'dozymetry': 'dozymetr_list',
            }
            return redirect(redirect_map.get(model_name, 'home'))

        except Exception as e:
            return HttpResponseBadRequest(f"Błąd podczas importu: {e}")